from openpyxl import load_workbook
from typing import List, Dict


class ExcelParser:
    """Parse Excel questionnaires to extract questions."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = load_workbook(filepath)
        self.ws = self.wb.active
    
    def parse(self) -> List[Dict]:
        """Extract questions from Excel."""
        questions = []
        
        # Skip header row (row 1)
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is None:
                continue
            
            question_num = row[0].value
            question_text = row[1].value
            
            if question_text:
                questions.append({
                    'number': question_num,
                    'question': question_text,
                    'row': row_idx,
                    'cell_ref': row[2].coordinate  # Answer column
                })
        
        return questions
