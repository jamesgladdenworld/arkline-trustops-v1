from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any
from pathlib import Path


class ExcelExporter:
    """Export reviewed responses back to Excel questionnaire."""
    
    def __init__(self, template_path: str, output_path: str):
        self.template_path = template_path
        self.output_path = output_path
        self.wb = load_workbook(template_path)
        self.ws = self.wb.active
    
    def export_responses(self, responses: List[Dict[str, Any]]):
        """Write approved responses to Excel."""
        for response in responses:
            row = response['row']
            cell_ref = response['cell_ref']
            answer = response['answer']
            
            cell = self.ws[cell_ref]
            cell.value = answer
            
            # Style the cell
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        
        # Save the workbook
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.output_path)
        print(f"✅ Exported to {self.output_path}")
    
    def export_summary(self, summary: Dict[str, Any], output_path: str):
        """Export a summary of the review process."""
        summary_text = f"""
Questionnaire Completion Summary
================================
Total Questions: {summary.get('total', 0)}
Approved: {summary.get('approved', 0)}
Rejected: {summary.get('rejected', 0)}
Edited: {summary.get('edited', 0)}
Completion Rate: {summary.get('completion_rate', 0):.1%}

Generated: {summary.get('timestamp', 'N/A')}
"""
        
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            f.write(summary_text)
        
        print(f"✅ Summary exported to {output_path}")
